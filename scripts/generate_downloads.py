# Build the 4 downloadable template files for the SustainSutra website.
# 2 PDF checklists (HTML -> Chrome headless print) + 2 Excel workbooks (openpyxl).
import os, subprocess, time

ROOT = r"D:\Application_Dev\sustainsutra-main\public\downloads"
os.makedirs(ROOT, exist_ok=True)

NAVY = "#0B0F0B"; GOLD = "#D4AF37"; OFF = "#F8FAFC"; DIM = "#A0AAB5"

# ------------------------------------------------------------------ PDF 1
CBAM_HTML = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>CBAM Exporter Readiness Checklist</title>
<style>
@page {{ size: A4; margin: 16mm 14mm; }}
* {{ box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #1a1a1a; font-size: 10.5pt; line-height: 1.5; }}
.cover {{ border-bottom: 4px solid {GOLD}; padding-bottom: 12px; margin-bottom: 18px; }}
.cover h1 {{ font-size: 21pt; margin: 2px 0 4px; color: {NAVY}; }}
.kicker {{ color: {GOLD}; font-weight: 700; letter-spacing: 2px; font-size: 9pt; text-transform: uppercase; }}
.cover p {{ color: #555; margin: 4px 0 0; font-size: 9.5pt; }}
h2 {{ font-size: 12.5pt; color: {NAVY}; border-left: 4px solid {GOLD}; padding-left: 8px; margin: 18px 0 8px; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 8px; }}
th {{ background: {NAVY}; color: #fff; text-align: left; padding: 6px 8px; font-size: 9pt; }}
td {{ border-bottom: 1px solid #e3e3e3; padding: 6px 8px; vertical-align: top; }}
td.n {{ width: 26px; color: {GOLD}; font-weight: 700; }}
td.box {{ width: 26px; border: 1.4px solid #bbb; }}
.foot {{ margin-top: 22px; padding-top: 10px; border-top: 2px solid {GOLD}; font-size: 8.5pt; color: #777; }}
</style></head><body>
<div class="cover">
  <div class="kicker">SustainSutra . Free Resource</div>
  <h1>CBAM Exporter Readiness Checklist</h1>
  <p>40-point self-assessment for exporters of CBAM-covered goods into the EU. Tick each item when the underlying evidence exists — not when it is planned.</p>
</div>

<h2>A. Product &amp; Scope (6)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Evidence / Notes</th></tr>
<tr><td class="box"></td><td class="n">1</td><td>Every exported product classified against EU CN codes; CBAM-covered vs not determined per product</td><td></td></tr>
<tr><td class="box"></td><td class="n">2</td><td>Downstream/derived products checked (some fall in scope even when the base product does not)</td><td></td></tr>
<tr><td class="box"></td><td class="n">3</td><td>Precursor goods entering the production process identified (their embedded emissions must be counted)</td><td></td></tr>
<tr><td class="box"></td><td class="n">4</td><td>EU customers identified: who acts as importer of record (declarant)?</td><td></td></tr>
<tr><td class="box"></td><td class="n">5</td><td>UK CBAM (Jan 2027) coverage assessed separately — scope differs from EU</td><td></td></tr>
<tr><td class="box"></td><td class="n">6</td><td>Annual tonnage and value per covered CN code compiled</td><td></td></tr>
</table>

<h2>B. Embedded Emissions Data (10)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Evidence / Notes</th></tr>
<tr><td class="box"></td><td class="n">7</td><td>Installation-level production process mapped (process flow with all emission sources)</td><td></td></tr>
<tr><td class="box"></td><td class="n">8</td><td>District heating / fuel / electricity consumption metered at installation level</td><td></td></tr>
<tr><td class="box"></td><td class="n">9</td><td>Product-level allocation rules defined (allocation to product vs co-products)</td><td></td></tr>
<tr><td class="box"></td><td class="n">10</td><td>Actual embedded emissions calculable (vs relying on default values)</td><td></td></tr>
<tr><td class="box"></td><td class="n">11</td><td>Emission factors cited with source and vintage for every fuel/material</td><td></td></tr>
<tr><td class="box"></td><td class="n">12</td><td>Electricity: grid factor source documented (country-specific, correct vintage)</td><td></td></tr>
<tr><td class="box"></td><td class="n">13</td><td>Precursor emissions included where required by the implementing regulation</td><td></td></tr>
<tr><td class="box"></td><td class="n">14</td><td>Internal QC: second person re-performs the calculation independently</td><td></td></tr>
<tr><td class="box"></td><td class="n">15</td><td>Data chain documented: source document to final number, reproducible</td><td></td></tr>
<tr><td class="box"></td><td class="n">16</td><td>Consistency checked between CBAM figures and company GHG inventory</td><td></td></tr>
</table>

<h2>C. Verification Readiness (8)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Evidence / Notes</th></tr>
<tr><td class="box"></td><td class="n">17</td><td>Evidence pack assembled (invoices, logbooks, meter records, calc files)</td><td></td></tr>
<tr><td class="box"></td><td class="n">18</td><td>Calculation files version-controlled with named owners</td><td></td></tr>
<tr><td class="box"></td><td class="n">19</td><td>Data collection SOP written (who records what, when, from where)</td><td></td></tr>
<tr><td class="box"></td><td class="n">20</td><td>Pre-verification gap assessment done (internal or external)</td><td></td></tr>
<tr><td class="box"></td><td class="n">21</td><td>Verification body identified / engaged</td><td></td></tr>
<tr><td class="box"></td><td class="n">22</td><td>Findings from gap assessment resolved or scheduled</td><td></td></tr>
<tr><td class="box"></td><td class="n">23</td><td>Authorised signatory identified for declarations</td><td></td></tr>
<tr><td class="box"></td><td class="n">24</td><td>Filing calendar known (quarterly/annual deadlines for the EU declarant)</td><td></td></tr>
</table>

<h2>D. Commercial &amp; Contract (6)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Evidence / Notes</th></tr>
<tr><td class="box"></td><td class="n">25</td><td>CBAM cost exposure quantified at current EU ETS price and phase-out share</td><td></td></tr>
<tr><td class="box"></td><td class="n">26</td><td>Forward contracts reviewed for CBAM liability allocation clauses</td><td></td></tr>
<tr><td class="box"></td><td class="n">27</td><td>Price-negotiation brief prepared (verified intensity vs defaults)</td><td></td></tr>
<tr><td class="box"></td><td class="n">28</td><td>Data-sharing terms with EU customers agreed (format, frequency)</td><td></td></tr>
<tr><td class="box"></td><td class="n">29</td><td>Confidentiality boundaries set on process data shared externally</td><td></td></tr>
<tr><td class="box"></td><td class="n">30</td><td>Alternative market options assessed (where EU exposure is marginal)</td><td></td></tr>
</table>

<h2>E. Systems &amp; Timeline (10)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Evidence / Notes</th></tr>
<tr><td class="box"></td><td class="n">31</td><td>Named internal owner for CBAM (single accountable person)</td><td></td></tr>
<tr><td class="box"></td><td class="n">32</td><td>Training delivered to the data-collection team</td><td></td></tr>
<tr><td class="box"></td><td class="n">33</td><td>Metering gaps listed with installation plan</td><td></td></tr>
<tr><td class="box"></td><td class="n">34</td><td>Calculation tool/system in place (not ad-hoc spreadsheets)</td><td></td></tr>
<tr><td class="box"></td><td class="n">35</td><td>Quarterly internal data review scheduled</td><td></td></tr>
<tr><td class="box"></td><td class="n">36</td><td>Registry/trader account status known (via your EU declarant)</td><td></td></tr>
<tr><td class="box"></td><td class="n">37</td><td>2026 declaration obligations understood (first definitive-year cycle)</td><td></td></tr>
<tr><td class="box"></td><td class="n">38</td><td>Phase-out schedule (to 2032) modelled for multi-year exposure</td><td></td></tr>
<tr><td class="box"></td><td class="n">39</td><td>Decarbonisation options screened (each tCO2e cut = certificate cost cut)</td><td></td></tr>
<tr><td class="box"></td><td class="n">40</td><td>Escalation path defined when data gaps appear mid-cycle</td><td></td></tr>
</table>

<div class="foot"><b>SustainSutra GreenTech LLP</b> . CBAM readiness, embedded emissions calculation, verification preparation . director@sustainsutra.in . sustainsutra.in<br>
This checklist is general guidance, not legal advice. Regulation details evolve — verify current requirements on the official EU CBAM portal.</div>
</body></html>"""

# ------------------------------------------------------------------ PDF 2
BRSR_HTML = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>BRSR Readiness Self-Assessment</title>
<style>
@page {{ size: A4; margin: 16mm 14mm; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #1a1a1a; font-size: 10.5pt; line-height: 1.5; }}
.cover {{ border-bottom: 4px solid {GOLD}; padding-bottom: 12px; margin-bottom: 18px; }}
.cover h1 {{ font-size: 21pt; margin: 2px 0 4px; color: {NAVY}; }}
.kicker {{ color: {GOLD}; font-weight: 700; letter-spacing: 2px; font-size: 9pt; text-transform: uppercase; }}
.cover p {{ color: #555; margin: 4px 0 0; font-size: 9.5pt; }}
h2 {{ font-size: 12.5pt; color: {NAVY}; border-left: 4px solid {GOLD}; padding-left: 8px; margin: 18px 0 8px; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 8px; }}
th {{ background: {NAVY}; color: #fff; text-align: left; padding: 6px 8px; font-size: 9pt; }}
td {{ border-bottom: 1px solid #e3e3e3; padding: 6px 8px; vertical-align: top; }}
td.n {{ width: 26px; color: {GOLD}; font-weight: 700; }}
td.box {{ width: 26px; border: 1.4px solid #bbb; }}
.foot {{ margin-top: 22px; padding-top: 10px; border-top: 2px solid {GOLD}; font-size: 8.5pt; color: #777; }}
</style></head><body>
<div class="cover">
  <div class="kicker">SustainSutra . Free Resource</div>
  <h1>BRSR Readiness Self-Assessment</h1>
  <p>Diagnostic across the BRSR disclosure architecture: data backbone, boundaries, principle-wise KPI coverage, leadership indicators and assurance readiness.</p>
</div>

<h2>A. Scope &amp; Governance (5)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Gap / Notes</th></tr>
<tr><td class="box"></td><td class="n">1</td><td>Listing status and BRSR applicability confirmed (mandatory vs voluntary filing)</td><td></td></tr>
<tr><td class="box"></td><td class="n">2</td><td>Reporting boundary decided: standalone vs consolidated (subsidiaries in/out defined)</td><td></td></tr>
<tr><td class="box"></td><td class="n">3</td><td>Board / CSR committee oversight of sustainability disclosure documented</td><td></td></tr>
<tr><td class="box"></td><td class="n">4</td><td>Named disclosure owner + data owners per principle</td><td></td></tr>
<tr><td class="box"></td><td class="n">5</td><td>Filing calendar with internal deadlines mapped (ahead of SEBI date)</td><td></td></tr>
</table>

<h2>B. Data Backbone (8)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Gap / Notes</th></tr>
<tr><td class="box"></td><td class="n">6</td><td>Energy consumption reconciles with energy audit / statutory filings</td><td></td></tr>
<tr><td class="box"></td><td class="n">7</td><td>Water withdrawal matches CGWA NOC / SPCB records where applicable</td><td></td></tr>
<tr><td class="box"></td><td class="n">8</td><td>GHG Scope 1 &amp; 2 computed on a documented method (factors cited)</td><td></td></tr>
<tr><td class="box"></td><td class="n">9</td><td>Waste categories tracked per rule (hazardous/e-waste/plastic/biomedical) with manifests</td><td></td></tr>
<tr><td class="box"></td><td class="n">10</td><td>Employee / worker data system covers the social KPIs (gender, turnover, training hours, PWD)</td><td></td></tr>
<tr><td class="box"></td><td class="n">11</td><td>Complaint/grievance register feeds the governance KPIs</td><td></td></tr>
<tr><td class="box"></td><td class="n">12</td><td>Value-chain data (suppliers, PP coverage) collected on a defined method</td><td></td></tr>
<tr><td class="box"></td><td class="n">13</td><td>Restatement policy: YoY method changes disclosed and prior year restated</td><td></td></tr>
</table>

<h2>C. Principle-Wise Coverage (9)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Gap / Notes</th></tr>
<tr><td class="box"></td><td class="n">14</td><td>P1 Ethics: code of conduct, awareness coverage % tracked</td><td></td></tr>
<tr><td class="box"></td><td class="n">15</td><td>P2 Sustainable &amp; safe products: product recalls, R&D spend on sustainability</td><td></td></tr>
<tr><td class="box"></td><td class="n">16</td><td>P3 Wellbeing: safety (LTIFR), parental leave, accessibility KPIs</td><td></td></tr>
<tr><td class="box"></td><td class="n">17</td><td>P4 Stakeholder engagement: mapped, consulted, feedback loop closed</td><td></td></tr>
<tr><td class="box"></td><td class="n">18</td><td>P5 Human rights: policy, due diligence, child/forced labour KPIs</td><td></td></tr>
<tr><td class="box"></td><td class="n">19</td><td>P6 Environment: all quantified sections (energy/water/GHG/waste/biodiversity)</td><td></td></tr>
<tr><td class="box"></td><td class="n">20</td><td>P7 Public policy: advocacy positions, PIC compliance</td><td></td></tr>
<tr><td class="box"></td><td class="n">21</td><td>P8 Inclusive growth: CSR spend, community project outcomes</td><td></td></tr>
<tr><td class="box"></td><td class="n">22</td><td>P9 Consumer: complaints, data privacy, advertising KPIs</td><td></td></tr>
</table>

<h2>D. Leadership Indicators (5)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Gap / Notes</th></tr>
<tr><td class="box"></td><td class="n">23</td><td>Leadership indicator strategy decided (which to pursue, which to skip honestly)</td><td></td></tr>
<tr><td class="box"></td><td class="n">24</td><td>Life-cycle assessment(s) planned or completed for key products</td><td></td></tr>
<tr><td class="box"></td><td class="n">25</td><td>Supply-chain ESG engagement programme running</td><td></td></tr>
<tr><td class="box"></td><td class="n">26</td><td>Third-party assurance scoped (limited vs reasonable)</td><td></td></tr>
<tr><td class="box"></td><td class="n">27</td><td>SDG mapping / alignment documented</td><td></td></tr>
</table>

<h2>E. Assurance Readiness (5)</h2>
<table><tr><th></th><th>#</th><th>Check</th><th>Gap / Notes</th></tr>
<tr><td class="box"></td><td class="n">28</td><td>Basis of preparation documented for every quantitative indicator</td><td></td></tr>
<tr><td class="box"></td><td class="n">29</td><td>Evidence pack per KPI (source docs traceable)</td><td></td></tr>
<tr><td class="box"></td><td class="n">30</td><td>Cross-document consistency: BRSR vs annual report vs website claims</td><td></td></tr>
<tr><td class="box"></td><td class="n">31</td><td>Assurance provider engaged; scope and timeline agreed</td><td></td></tr>
<tr><td class="box"></td><td class="n">32</td><td>Internal mock-review against the assurance checklist completed</td><td></td></tr>
</table>

<div class="foot"><b>SustainSutra GreenTech LLP</b> . BRSR readiness, reporting and assurance preparation . director@sustainsutra.in . sustainsutra.in<br>
This assessment is general guidance. Refer to the current SEBI BRSR circular and format for binding requirements.</div>
</body></html>"""

CHROME = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

def html_to_pdf(html, pdf_name):
    html_path = os.path.join(ROOT, pdf_name.replace(".pdf", ".html"))
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    out = os.path.join(ROOT, pdf_name)
    cmd = [CHROME, "--headless", "--disable-gpu", "--no-sandbox",
           f"--print-to-pdf={out}", "--no-pdf-header-footer", html_path]
    subprocess.run(cmd, check=True, capture_output=True, timeout=90)
    print("saved", out, os.path.getsize(out), "bytes")

html_to_pdf(CBAM_HTML, "SustainSutra_CBAM_Exporter_Readiness_Checklist.pdf")
html_to_pdf(BRSR_HTML, "SustainSutra_BRSR_Readiness_Checklist.pdf")

# ------------------------------------------------------------------ Excel 1: Consent & Compliance Calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_HEX = "0B0F0B"; GOLD_HEX = "D4AF37"; LIGHT = "F4F1E8"

hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=NAVY_HEX)
sub_fill = PatternFill("solid", fgColor="1E2A38")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

wb = Workbook()

# --- Sheet 1: Consent tracker
ws = wb.active; ws.title = "Consent Tracker"
ws.append(["S.No", "Consent / Approval", "Category (Red/Orange/Green)", "Consent No.", "Issue Date", "Valid Upto",
           "Days Left", "Status", "Renewal Application Due (apply 60d before)", "Owner", "Notes"])
style_header(ws, 1, 11)
consents = [
    ["Consent to Operate (CTO)", "Red", "", "", "", "", "", "", "", "Renew 60 days before expiry; SPCB may take 30-60d"],
    ["Consent to Operate (CTO)", "Orange", "", "", "", "", "", "", "", ""],
    ["Consent to Establish (CTE)", "", "", "", "", "", "", "", "", "If expanding / new product line"],
    ["Hazardous Waste Authorisation", "", "", "", "", "", "", "", "", "Form 1 + agreement with TSDF"],
    ["Authorisation - PWM / E-Waste / Battery", "", "", "", "", "", "", "", "", "As applicable per rule"],
    ["CGWA NOC (groundwater)", "", "", "", "", "", "", "", "", "If borewells in use"],
]
for i, row in enumerate(consents, start=2):
    ws.append([i-1] + row)
ws.append([])
ws.append(["STATUS FORMULA: Days Left = Valid Upto - TODAY(); if <60 show AMBER, if <30 RED. Enter dates as real dates."])
for c, w in enumerate([6, 26, 16, 14, 12, 12, 10, 12, 24, 12, 30], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# --- Sheet 2: Annual compliance calendar
ws2 = wb.create_sheet("Annual Calendar")
ws2.append(["Month", "Obligation", "Frequency", "Due Date", "Authority / Portal", "Status", "Evidence Link", "Notes"])
style_header(ws2, 1, 8)
cal = [
    ["Jan", "Monthly lab monitoring (air/stack/noise) sample plan", "Monthly", "Varies", "NABL lab + SPCB", "", "", "As per consent conditions"],
    ["Feb", "Monthly lab monitoring", "Monthly", "Varies", "NABL lab + SPCB", "", "", ""],
    ["Mar", "Monthly lab monitoring; FY data compilation begins", "Monthly", "Varies", "NABL lab + SPCB", "", "", "Close FY books for env data"],
    ["Apr", "Annual return - Hazardous Waste (if FY = Apr-Mar)", "Annual", "30 Jun", "SPCB / CPCB portal", "", "", "Form 4 / annual report"],
    ["May", "Monthly lab monitoring", "Monthly", "Varies", "", "", "", ""],
    ["Jun", "Hazardous Waste annual return (Form 4)", "Annual", "30 Jun", "SPCB", "", "", "COMMON MISS"],
    ["Jul", "Quarterly report Q1 (where applicable)", "Quarterly", "Varies", "SPCB", "", "", ""],
    ["Aug", "Monthly lab monitoring", "Monthly", "Varies", "", "", "", ""],
    ["Sep", "ENVIRONMENTAL STATEMENT (Form V)", "Annual", "30 Sep", "SPCB", "", "", "COMMON MISS - mandatory u/s EPA Rules"],
    ["Oct", "PWM / E-Waste / Battery EPR returns (as applicable)", "Annual/Quarterly", "Varies", "CPCB portals", "", "", ""],
    ["Nov", "Quarterly report Q3", "Quarterly", "Varies", "SPCB", "", "", ""],
    ["Dec", "OCEMS connectivity check + data reconciliation", "Quarterly", "Varies", "CPCB OCEMS", "", "", ""],
]
for row in cal:
    ws2.append(row)
for c, w in enumerate([8, 40, 13, 12, 18, 10, 14, 30], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# --- Sheet 3: Lab monitoring schedule
ws3 = wb.create_sheet("Lab Monitoring")
ws3.append(["Parameter Group", "Typical Frequency", "Consent Reference", "Last Done", "Next Due", "Lab Used", "Result Summary", "Action if Non-compliant"])
style_header(ws3, 1, 8)
mon = [
    ["Stack emissions (PM, SO2, NOx, process-specific)", "Monthly / Quarterly per consent", "", "", "", "", "", "Re-test + corrective action report to SPCB"],
    ["Ambient air quality (PM10, PM2.5, SO2, NO2)", "Monthly / Quarterly", "", "", "", "", "", ""],
    ["Effluent quality (pH, BOD, COD, TSS, sector params)", "Monthly", "", "", "", "", "", ""],
    ["Noise (boundary, day/night)", "Quarterly / Half-yearly", "", "", "", "", "", ""],
    ["Hazardous waste analysis (where required)", "Annual", "", "", "", "", "", ""],
    ["Groundwater / soil (if in consent)", "Half-yearly / Annual", "", "", "", "", "", ""],
]
for row in mon:
    ws3.append(row)
for c, w in enumerate([38, 22, 16, 12, 12, 14, 20, 26], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

wb.save(os.path.join(ROOT, "SustainSutra_Consent_Compliance_Calendar.xlsx"))
print("saved Excel: consent calendar")

# ------------------------------------------------------------------ Excel 2: GHG data collection template
wb2 = Workbook()
wsA = wb2.active; wsA.title = "README"
wsA["A1"] = "SustainSutra - GHG Inventory Data Collection Template"
wsA["A1"].font = Font(bold=True, size=14, color=NAVY_HEX)
notes = [
    "",
    "Aligned with ISO 14064-1:2018 categorisation.",
    "Rules:",
    "1. One row per data point. Never overwrite a number - correct it in the file with a change note.",
    "2. Every activity figure must trace to a source document (invoice, logbook, meter export) - record it.",
    "3. Every emission factor needs source + vintage. If judgement was used, write why in the notes.",
    "4. Mark estimates clearly in the Data Quality column. Estimates are acceptable; undocumented estimates fail verification.",
    "5. Data owner = the person accountable for the number, not the person who typed it.",
]
for i, n in enumerate(notes, start=2):
    wsA[f"A{i}"] = n
wsA.column_dimensions["A"].width = 110

cats = [
    ("Scope 1 - Stationary Combustion", [
        "Fuel type (coal/HSD/FO/ng/LPG/biomass)", "Quantity consumed", "Unit (t / kL / Sm3)", "NCV used", "EF source", "CO2e factor (t/unit)", "Source document ref", "Data owner", "Data quality (measured/estimated)", "Notes"]),
    ("Scope 1 - Mobile Combustion", [
        "Vehicle/equipment", "Fuel type", "Quantity", "Unit", "EF source", "CO2e factor", "Source document ref", "Data owner", "Data quality", "Notes"]),
    ("Scope 1 - Process Emissions", [
        "Process", "Feedstock/product quantity", "Unit", "Stoichiometric or measured factor", "EF source", "CO2e factor", "Source document ref", "Data owner", "Data quality", "Notes"]),
    ("Scope 2 - Purchased Electricity", [
        "Meter / facility", "kWh consumed", "Grid factor source (CEA vintage)", "CO2e factor (tCO2/MWh)", "Source document ref", "Data owner", "Data quality", "Notes"]),
    ("Scope 3 - Category 1 Purchased Goods", [
        "Material", "Quantity", "Unit", "Supplier-specific data? (Y/N)", "EF source (spend/average/hybrid)", "CO2e factor", "Source document ref", "Data owner", "Data quality", "Notes"]),
    ("Scope 3 - Other Relevant Categories", [
        "Category (2-15)", "Activity description", "Activity data", "Unit", "Method (spend/average/supplier)", "CO2e factor", "Source document ref", "Data owner", "Data quality", "Notes"]),
]
for title, cols in cats:
    sh = wb2.create_sheet(title[:31])
    sh.append(cols)
    style_header(sh, 1, len(cols))
    for c, w in enumerate([26, 16, 14, 16, 24, 16, 22, 14, 20, 24], 1):
        try: sh.column_dimensions[get_column_letter(c)].width = min(w, 26)
        except Exception: pass
    for r in range(2, 32):
        sh.cell(row=r, column=1).border = border
wb2.save(os.path.join(ROOT, "SustainSutra_GHG_Data_Collection_Template.xlsx"))
print("saved Excel: GHG template")
print("ALL DOWNLOADS DONE")
